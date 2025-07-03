# import
import os
import time
import shutil
import xml.etree.ElementTree as ET
import pandas as pd
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook, Workbook

#Cấu hình web
def open_browser(thu_muc_tai_hoa_don):
    os.makedirs(thu_muc_tai_hoa_don, exist_ok=True)
    options = Options()
    options.add_experimental_option("prefs", {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "download.default_directory": thu_muc_tai_hoa_don,
        "plugins.always_open_pdf_externally": True,
        "profile.default_content_settings.popups": 0,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1
    })
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    driver = webdriver.Chrome(service=Service(), options=options)
    return driver, WebDriverWait(driver, 10)



def tra_cuu_fpt(driver, wait, ma_so_thue, ma_tra_cuu):
    driver.get("https://tracuuhoadon.fpt.com.vn/search.html")
    mst_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='MST bên bán']")))
    mst_input.clear()
    mst_input.send_keys(ma_so_thue.strip().replace("'", ""))

    mtc_input = driver.find_element(By.XPATH, "//input[@placeholder='Mã tra cứu hóa đơn']")
    mtc_input.clear()
    mtc_input.send_keys(ma_tra_cuu.strip())

    btn = driver.find_element(By.XPATH, "//button[contains(@class, 'webix_button') and contains(text(), 'Tra cứu')]")
    btn.click()

def tra_cuu_meinvoice(driver, wait, ma_tra_cuu):
    driver.get("https://www.meinvoice.vn/tra-cuu/")
    mtc_input = wait.until(EC.presence_of_element_located((By.NAME, "txtCode")))
    driver.execute_script("""const header = document.querySelector('.top-header'); if (header) header.style.display = 'none';""")
    mtc_input.clear()
    mtc_input.send_keys(ma_tra_cuu.strip())

    btn = wait.until(EC.element_to_be_clickable((By.ID, "btnSearchInvoice")))
    btn.click()

def tra_cuu_ehoadon(driver, wait, ma_tra_cuu):
    driver.get("https://van.ehoadon.vn/TCHD?MTC")
    mtc_input = wait.until(EC.presence_of_element_located((By.ID, "txtInvoiceCode")))
    mtc_input.clear()
    mtc_input.send_keys(ma_tra_cuu)
    driver.find_element(By.CLASS_NAME, "btnSearch").click()

# Tải ảnh dạng file XML
def tai_xml_fpt(driver, wait):
    btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[span[contains(@class, 'mdi-xml')] and contains(text(), 'Tải XML')]")))
    btn.click()

def tai_xml_meinvoice(driver, wait):
    btn_menu = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "download")))
    driver.execute_script("arguments[0].click();", btn_menu)
    btn_xml = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "txt-download-xml")))
    driver.execute_script("arguments[0].click();", btn_xml)

def tai_xml_ehoadon(driver, wait):
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "frameViewInvoice")))
    btn = wait.until(EC.presence_of_element_located((By.ID, "btnDownload")))
    ActionChains(driver).move_to_element(btn).perform()
    driver.execute_script("document.querySelector('#divDownloads .dropdown-menu').style.display='block';")
    btn_xml = wait.until(EC.element_to_be_clickable((By.ID, "LinkDownXML")))
    btn_xml.click()
    driver.switch_to.default_content()

# Tra cứu và tải
def tra_cuu_hoa_don(driver, wait, ma_so_thue, ma_tra_cuu, url):
    try:
        if "fpt.com.vn" in url:
            tra_cuu_fpt(driver, wait, ma_so_thue, ma_tra_cuu)
        elif "meinvoice.vn" in url:
            tra_cuu_meinvoice(driver, wait, ma_tra_cuu)
        elif "ehoadon.vn" in url:
            tra_cuu_ehoadon(driver, wait, ma_tra_cuu)
        else:
            print(f"Chưa hỗ trợ trang: {url}")
            return False

        wait.until(EC.presence_of_element_located((By.XPATH, "//body")))
        return True

    except Exception as e:
        print(f"Lỗi khi tra cứu: {e}")
        return False

def tai_file_xml(driver, wait, thu_muc_tai_hoa_don, url, ma_tra_cuu):
    try:
        if "fpt.com.vn" in url:
            tai_xml_fpt(driver, wait)
        elif "meinvoice.vn" in url:
            tai_xml_meinvoice(driver, wait)
        elif "ehoadon.vn" in url:
            tai_xml_ehoadon(driver, wait)
        else:
            print(f"Chưa hỗ trợ tải từ: {url}")
            return None

        domain_folder = os.path.join(thu_muc_tai_hoa_don, urlparse(url).netloc.replace("www.", ""))
        os.makedirs(domain_folder, exist_ok=True)

        for _ in range(10):
            files = os.listdir(thu_muc_tai_hoa_don)
            for file in files:
                if file.endswith(".xml"):
                    dest = os.path.join(domain_folder, f"{ma_tra_cuu}.xml")
                    shutil.move(os.path.join(thu_muc_tai_hoa_don, file), dest)
                    return dest
            time.sleep(1)

    except Exception as e:
        print(f"Lỗi tải file: {e}")
    return None

# Xử lí đọc file UML
def read_invoice_xml(xml_file_path):
    try:
        tree = ET.parse(xml_file_path)
        root = tree.getroot()

        for tag in [".//HDon/DLHDon", ".//DLHDon", ".//TDiep", ".//Invoice"]:
            node = root.find(tag)
            if node is not None:
                break
        else:
            print(f"Không tìm thấy dữ liệu chính: {xml_file_path}")
            return None

        def find(path):
            current = node
            for part in path.split("/"):
                current = current.find(part) if current is not None else None
            return current.text if current is not None else None

        stk_ban = find("NDHDon/NBan/STKNHang")
        if not stk_ban:
            for el in node.findall(".//NBan/TTKhac/TTin"):
                if el.findtext("TTruong") == "SellerBankAccount":
                    stk_ban = el.findtext("DLieu")
                    break

        return {
            'Số hóa đơn': find("TTChung/SHDon"),
            'Đơn vị bán hàng': find("NDHDon/NBan/Ten"),
            'Mã số thuế bán': find("NDHDon/NBan/MST"),
            'Địa chỉ bán': find("NDHDon/NBan/DChi"),
            'Số tài khoản bán': stk_ban,
            'Tên người mua hàng': find("NDHDon/NMua/Ten"),
            'Địa chỉ mua': find("NDHDon/NMua/DChi"),
            'MST mua hàng': find("NDHDon/NMua/MST"),
        }

    except Exception as e:
        print(f"Lỗi đọc XML {xml_file_path}: {e}")
        return None

# Ghi ra dạng file excel
def append_to_excel(filepath, row_data):
    if not os.path.isfile(filepath):
        wb = Workbook()
        ws = wb.active
        ws.append(["STT", "Mã số thuế", "Mã tra cứu", "URL", "Số hóa đơn", "Đơn vị bán hàng", "Mã số thuế bán", "Địa chỉ bán", "Số tài khoản bán", "Tên người mua hàng", "Địa chỉ mua", "MST mua hàng"])
        wb.save(filepath)

    wb = load_workbook(filepath)
    ws = wb.active
    ws.append(row_data)
    wb.save(filepath)

# chương trinh chính 
def main():
    input_file = "input.xlsx"
    output_file = "hoa_don.xlsx"
    thu_muc = os.path.join(os.getcwd(), "InvoiceData")
    driver, wait = open_browser(thu_muc)

    df = pd.read_excel(input_file, dtype=str)
    for idx, row in df.iterrows():
        stt = idx + 1
        mst = str(row.get("Mã số thuế", "")).strip()
        mtc = str(row.get("Mã tra cứu", "")).strip()
        url = str(row.get("URL", "")).strip()

        if not url or not mtc:
            continue

        print(f"\nTra cứu {mtc} tại {url}")
        if tra_cuu_hoa_don(driver, wait, mst, mtc, url):
            xml_path = tai_file_xml(driver, wait, thu_muc, url, mtc)
            parsed = read_invoice_xml(xml_path) if xml_path else None

            if parsed:
                row_data = [stt, mst, mtc, url] + list(parsed.values())
            else:
                row_data = [stt, mst, mtc, url] + [""] * 9

            append_to_excel(output_file, row_data)

    driver.quit()
    print(f"Hoàn tất. Kết quả lưu tại: {output_file}")

if __name__ == "__main__":
    main()
